from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\junyoung\source\repos\minigame-project-suk558165\Assets\Data"

HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
KR_FILL     = PatternFill("solid", start_color="4472C4")
KR_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left", vertical="center")
thin        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

def h(cell, fill, font):
    cell.fill = fill; cell.font = font
    cell.alignment = CENTER; cell.border = BORDER

def d(cell, alt=False):
    cell.font = DATA_FONT; cell.alignment = LEFT; cell.border = BORDER
    if alt: cell.fill = ALT_FILL

def auto_w(ws, extra=3):
    for col in ws.columns:
        L = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[L].width = w + extra

WEAPON_KR = ["id","무기타입","데미지","공격쿨다운(초)","스윙각도(도)","스윙지속시간(초)","투사체속도","가격"]
WEAPON_EN = ["id","weaponType","damage","attackCooldown","swingAngle","swingDuration","projectileSpeed","price"]
WEAPON_DATA = [
    # id          | type    | dmg | cool | angle | swing | proj |price
    ["knife",   "Melee",  18, 0.30,  90, 0.30,  0,  80],  # 칼
    ["bow",     "Ranged", 15, 0.80,   0, 0,    12, 150],  # 활
    ["crossbow","Ranged", 28, 1.40,   0, 0,    20, 200],  # 석궁
    ["spear",   "Melee",  22, 0.70,  60, 0.50,  0, 160],  # 창
    ["staff",   "Ranged", 20, 1.00,   0, 0,     9, 220],  # 마법봉
    ["scythe",  "Melee",  16, 0.50, 160, 0.45,  0, 140],  # 낫
    ["gauntlet","Melee",  10, 0.20,  70, 0.20,  0, 120],  # 장갑
    ["hammer",  "Melee",  38, 1.60, 100, 0.90,  0, 210],  # 망치
    ["hoe",     "Melee",  12, 0.40, 110, 0.40,  0,  90],  # 호미
]

ACC_KR = ["id","최대HP보너스","데미지보너스","이동속도보너스","점프보너스",
          "크리티컬확률","크리티컬데미지","공격속도보너스","피해감소",
          "받는피해배율","주는피해배율",
          "대쉬횟수추가","대쉬범위보너스",
          "회피율","골드드랍보너스",
          "화살갈래수","화살데미지배율","관통수","가격"]
ACC_EN = ["id","maxHpBonus","damageBonus","speedBonus","jumpBonus",
          "criticalChance","criticalDamage","attackSpeedBonus","damageReduction",
          "damageReceivedMult","damageDealtMult",
          "dashCountBonus","dashRangeBonus",
          "evasionRate","goldDropBonus",
          "arrowCount","arrowDamageMult","penetrationCount","price"]

ACC_DATA = [
    ["dash_boost",    0,   0, 0.30, 0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 120],
    ["move_speed",    0,   0, 0.20, 0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 100],
    ["critical",      0,   0, 0,    0, 0.15, 0.50,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 150],
    ["attack_speed",  0,   0, 0,    0,    0,    0, 0.25,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 130],
    ["max_hp",       30,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 120],
    ["gold_drop",     0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0,    0, 0.30, 0,   0, 0, 140],
    ["damage_reduce", 0,   0, 0,    0,    0,    0,    0, 0.15,    0,    0, 0, 0,    0,    0, 0,   0, 0, 160],
    ["risk_reward",   0,   0, 0,    0,    0,    0,    0,    0, 0.30, 0.50, 0, 0,    0,    0, 0,   0, 0, 180],
    ["jump_power",    0,   0, 0, 0.30,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 100],
    ["dash_count",    0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 1, 0,    0,    0, 0,   0, 0, 150],
    ["evasion",       0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0, 0.10,    0, 0,   0, 0, 170],
    ["triple_arrow",  0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 3, 0.5, 0, 200],
    ["dash_range",    0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0.40, 0,    0, 0,   0, 0, 130],
    ["penetration",   0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 2, 220],
    ["sell_item",     0,   0, 0,    0,    0,    0,    0,    0,    0,    0, 0, 0,    0,    0, 0,   0, 0, 500],
]

wb1 = Workbook()
ws_w = wb1.active; ws_w.title = "Weapon"; ws_w.freeze_panes = "A3"
for c, v in enumerate(WEAPON_KR, 1): h(ws_w.cell(1, c, v), KR_FILL, KR_FONT)
for c, v in enumerate(WEAPON_EN, 1): h(ws_w.cell(2, c, v), HEADER_FILL, HEADER_FONT)
for r, row in enumerate(WEAPON_DATA, 3):
    for c, val in enumerate(row, 1): d(ws_w.cell(r, c, val), r%2==0)
auto_w(ws_w)

ws_a = wb1.create_sheet("Accessory"); ws_a.freeze_panes = "A3"
for c, v in enumerate(ACC_KR, 1): h(ws_a.cell(1, c, v), KR_FILL, KR_FONT)
for c, v in enumerate(ACC_EN, 1): h(ws_a.cell(2, c, v), HEADER_FILL, HEADER_FONT)
for r, row in enumerate(ACC_DATA, 3):
    for c, val in enumerate(row, 1): d(ws_a.cell(r, c, val), r%2==0)
auto_w(ws_a)

wb1.save(OUTPUT_DIR + r"\ItemNumericData.xlsx")
print("ItemNumericData.xlsx saved")

STR_KR = ["id", "Key", "한국어", "English"]

WEAPON_STR = [
    ["knife",    "weapon_knife_name",    "칼",     "Knife"],
    ["knife",    "weapon_knife_desc",    "빠르고 가벼운 근접 무기.",             "A fast and light melee weapon."],
    ["bow",      "weapon_bow_name",      "활",     "Bow"],
    ["bow",      "weapon_bow_desc",      "원거리 화살을 발사하는 활.",           "A bow that fires arrows from a distance."],
    ["crossbow", "weapon_crossbow_name", "석궁",   "Crossbow"],
    ["crossbow", "weapon_crossbow_desc", "강력하지만 재장전이 느린 원거리 무기.", "A powerful ranged weapon with a slow reload."],
    ["spear",    "weapon_spear_name",    "창",     "Spear"],
    ["spear",    "weapon_spear_desc",    "긴 사거리의 찌르기 근접 무기.",        "A thrusting melee weapon with long reach."],
    ["staff",    "weapon_staff_name",    "마법봉", "Magic Staff"],
    ["staff",    "weapon_staff_desc",    "마법 에너지를 발사하는 봉.",           "A staff that fires magical energy."],
    ["scythe",   "weapon_scythe_name",   "낫",     "Scythe"],
    ["scythe",   "weapon_scythe_desc",   "넓은 범위를 베는 근접 무기.",          "A melee weapon that slashes in a wide arc."],
    ["gauntlet", "weapon_gauntlet_name", "장갑",   "Gauntlet"],
    ["gauntlet", "weapon_gauntlet_desc", "주먹으로 빠르게 공격하는 무기.",       "A weapon that strikes fast with bare fists."],
    ["hammer",   "weapon_hammer_name",   "망치",   "Hammer"],
    ["hammer",   "weapon_hammer_desc",   "느리지만 강력한 타격을 주는 무기.",    "A slow but devastatingly powerful weapon."],
    ["hoe",      "weapon_hoe_name",      "호미",   "Hoe"],
    ["hoe",      "weapon_hoe_desc",      "농사 도구이지만 꽤 쓸만하다.",         "A farming tool that works surprisingly well."],
]

ACC_STR = [
    ["dash_boost",    "acc_dash_boost_name",    "질주의 부적",       "Sprint Charm"],
    ["dash_boost",    "acc_dash_boost_desc",    "이동속도가 크게 증가한다.", "Greatly increases movement speed."],
    ["move_speed",    "acc_move_speed_name",    "바람의 신발",        "Wind Boots"],
    ["move_speed",    "acc_move_speed_desc",    "이동속도가 증가한다.", "Increases movement speed."],
    ["critical",      "acc_critical_name",      "예리한 눈",          "Sharp Eye"],
    ["critical",      "acc_critical_desc",      "크리티컬 확률과 데미지가 증가한다.", "Increases critical chance and damage."],
    ["attack_speed",  "acc_attack_speed_name",  "빠른 손목",          "Quick Wrist"],
    ["attack_speed",  "acc_attack_speed_desc",  "공격 속도가 증가한다.", "Increases attack speed."],
    ["max_hp",        "acc_max_hp_name",        "강인한 심장",         "Sturdy Heart"],
    ["max_hp",        "acc_max_hp_desc",        "최대 체력이 증가한다.", "Increases maximum HP."],
    ["gold_drop",     "acc_gold_drop_name",     "황금 손",            "Golden Hand"],
    ["gold_drop",     "acc_gold_drop_desc",     "적 처치 시 골드 드랍량이 증가한다.", "Enemies drop more gold on defeat."],
    ["damage_reduce", "acc_damage_reduce_name", "철갑 방패",          "Iron Shield"],
    ["damage_reduce", "acc_damage_reduce_desc", "받는 피해가 감소한다.", "Reduces incoming damage."],
    ["risk_reward",   "acc_risk_reward_name",   "광전사의 각오",       "Berserker's Resolve"],
    ["risk_reward",   "acc_risk_reward_desc",   "받는 피해가 증가하지만 주는 피해도 크게 증가한다.", "Take more damage, but deal much more in return."],
    ["jump_power",    "acc_jump_power_name",    "스프링 부츠",         "Spring Boots"],
    ["jump_power",    "acc_jump_power_desc",    "점프력이 강화된다.",   "Enhances jump power."],
    ["dash_count",    "acc_dash_count_name",    "순간이동 부적",       "Blink Charm"],
    ["dash_count",    "acc_dash_count_desc",    "대쉬 횟수가 1회 추가된다.", "Grants one additional dash charge."],
    ["evasion",       "acc_evasion_name",       "그림자 망토",         "Shadow Cloak"],
    ["evasion",       "acc_evasion_desc",       "공격을 회피할 확률이 생긴다.", "Grants a chance to evade attacks."],
    ["triple_arrow",  "acc_triple_arrow_name",  "삼지창 화살촉",       "Triple Arrowhead"],
    ["triple_arrow",  "acc_triple_arrow_desc",  "화살이 3갈래로 나뉘지만 각 화살의 데미지가 감소한다.", "Arrows split into 3, but each deals less damage."],
    ["dash_range",    "acc_dash_range_name",    "질풍 발걸음",         "Gale Step"],
    ["dash_range",    "acc_dash_range_desc",    "대쉬 이동 거리가 증가한다.", "Increases dash travel distance."],
    ["penetration",   "acc_penetration_name",   "관통의 화살촉",       "Piercing Arrowhead"],
    ["penetration",   "acc_penetration_desc",   "화살이 적을 관통하여 뒤의 적도 공격한다.", "Arrows pierce through enemies, hitting those behind."],
    ["sell_item",     "acc_sell_item_name",     "상인의 보석",         "Merchant's Gem"],
    ["sell_item",     "acc_sell_item_desc",     "아무런 효과가 없지만 비싸게 팔 수 있다.", "Has no effect, but can be sold for a high price."],
]

wb2 = Workbook()

ws_ws = wb2.active; ws_ws.title = "Weapon"; ws_ws.freeze_panes = "A2"
for c, v in enumerate(STR_KR, 1): h(ws_ws.cell(1, c, v), KR_FILL, KR_FONT)
for r, row in enumerate(WEAPON_STR, 2):
    for c, val in enumerate(row, 1): d(ws_ws.cell(r, c, val), r%2==0)
auto_w(ws_ws)

ws_as = wb2.create_sheet("Accessory"); ws_as.freeze_panes = "A2"
for c, v in enumerate(STR_KR, 1): h(ws_as.cell(1, c, v), KR_FILL, KR_FONT)
for r, row in enumerate(ACC_STR, 2):
    for c, val in enumerate(row, 1): d(ws_as.cell(r, c, val), r%2==0)
auto_w(ws_as)

wb2.save(OUTPUT_DIR + r"\ItemStringData.xlsx")
print("ItemStringData.xlsx saved")

# ── CSV 출력 (Unity Editor 스크립트용) ───────────────────────────────────────
import csv, os

CSV_DIR = OUTPUT_DIR + r"\CSV"
os.makedirs(CSV_DIR, exist_ok=True)

def write_csv(filename, headers, rows):
    path = os.path.join(CSV_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows([headers] + rows)
    print(f"CSV saved: {path}")

write_csv("ItemNumericData_Weapon.csv",    WEAPON_EN, WEAPON_DATA)
write_csv("ItemNumericData_Accessory.csv", ACC_EN,    ACC_DATA)
write_csv("ItemStringData_Weapon.csv",
          ["id", "Key", "한국어", "English"],
          [[r[0], r[1], r[2], r[3]] for r in WEAPON_STR])
write_csv("ItemStringData_Accessory.csv",
          ["id", "Key", "한국어", "English"],
          [[r[0], r[1], r[2], r[3]] for r in ACC_STR])
